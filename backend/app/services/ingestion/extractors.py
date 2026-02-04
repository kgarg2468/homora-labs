import io
from pathlib import Path
from typing import AsyncGenerator

import fitz  # PyMuPDF
from docx import Document as DocxDocument
from openpyxl import load_workbook

from app.models.document import FileType


class PageContent:
    def __init__(self, text: str, page_number: int, has_images: bool = False, images: list = None):
        self.text = text
        self.page_number = page_number
        self.has_images = has_images
        self.images = images or []


async def extract_text_from_file(
    file_path: str, file_type: FileType
) -> AsyncGenerator[PageContent, None]:
    """Extract text content from various file types, yielding page by page."""
    path = Path(file_path)

    if file_type == FileType.pdf:
        async for page in extract_pdf(path):
            yield page
    elif file_type == FileType.docx:
        async for page in extract_docx(path):
            yield page
    elif file_type == FileType.xlsx:
        async for page in extract_xlsx(path):
            yield page
    elif file_type == FileType.image:
        yield PageContent(text="", page_number=1, has_images=True, images=[path])


async def extract_pdf(path: Path) -> AsyncGenerator[PageContent, None]:
    """Extract text and images from PDF file."""
    doc = fitz.open(str(path))

    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text("text")

        # Check for images
        images = []
        image_list = page.get_images()
        has_images = len(image_list) > 0

        # If page has little text but has images, mark for OCR
        if len(text.strip()) < 50 and has_images:
            # Extract images for OCR processing
            for img_idx, img in enumerate(image_list):
                xref = img[0]
                base_image = doc.extract_image(xref)
                if base_image:
                    images.append({
                        "data": base_image["image"],
                        "ext": base_image["ext"],
                    })

        yield PageContent(
            text=text,
            page_number=page_num + 1,
            has_images=has_images,
            images=images,
        )

    doc.close()


async def extract_docx(path: Path) -> AsyncGenerator[PageContent, None]:
    """Extract text from DOCX file."""
    doc = DocxDocument(str(path))

    all_text = []
    for para in doc.paragraphs:
        all_text.append(para.text)

    # DOCX doesn't have clear page boundaries, treat as single page
    yield PageContent(text="\n".join(all_text), page_number=1)


async def extract_xlsx(path: Path) -> AsyncGenerator[PageContent, None]:
    """Extract data from XLSX file and convert to markdown tables."""
    wb = load_workbook(str(path), data_only=True)

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        sheet = wb[sheet_name]
        markdown_table = convert_sheet_to_markdown(sheet, sheet_name)
        yield PageContent(text=markdown_table, page_number=sheet_idx + 1)


def convert_sheet_to_markdown(sheet, sheet_name: str) -> str:
    """Convert an Excel sheet to markdown table format."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return f"## {sheet_name}\n\n(Empty sheet)"

    # Filter out completely empty rows
    rows = [row for row in rows if any(cell is not None for cell in row)]
    if not rows:
        return f"## {sheet_name}\n\n(Empty sheet)"

    # Find max column width
    max_cols = max(len(row) for row in rows)

    # Pad rows to equal length
    rows = [list(row) + [None] * (max_cols - len(row)) for row in rows]

    # Build markdown table
    lines = [f"## {sheet_name}\n"]

    # Header row
    header = rows[0]
    header_cells = [str(cell) if cell is not None else "" for cell in header]
    lines.append("| " + " | ".join(header_cells) + " |")

    # Separator
    lines.append("| " + " | ".join(["---"] * max_cols) + " |")

    # Data rows
    for row in rows[1:]:
        cells = [str(cell) if cell is not None else "" for cell in row]
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def get_page_count(file_path: str, file_type: FileType) -> int | None:
    """Get page count for a file."""
    path = Path(file_path)

    if file_type == FileType.pdf:
        doc = fitz.open(str(path))
        count = len(doc)
        doc.close()
        return count
    elif file_type == FileType.xlsx:
        wb = load_workbook(str(path), data_only=True)
        return len(wb.sheetnames)
    elif file_type == FileType.docx:
        return 1  # DOCX doesn't have clear page boundaries
    elif file_type == FileType.image:
        return 1

    return None
